from table_reader import fetch_expected_specs_in_
import openpyxl as excel
import os
import uuid
expected_path = '/home/user/pytemp'
expected_filename = 'pies.xlsx'
maximum = {
    'cols': 10,
    'rows': 10
}
starting = {
    'col': 2,
    'row': 2
}
special_pies = ['jeremiah\'s', 'thomas\'s']

# resolve working directory
if os.getcwd() is not expected_path:
    os.chdir(expected_path)

# get reference to sheet
wb = excel.load_workbook(expected_filename)
c_sheet = wb['cherry']

# establish iteration limits
row_count = c_sheet.max_row \
    if c_sheet.max_row < maximum['rows'] \
    else maximum['rows']
col_count = c_sheet.max_column \
    if c_sheet.max_column < maximum['cols'] \
    else maximum['cols']

# extract data from the sheet
decision_col = 4
pie_data = {}
for active_row in range(starting['row'], row_count + 1):
    pie = c_sheet.cell(row=active_row,
                       column=decision_col).value
    if pie not in special_pies:
        continue
    for active_column in range(starting['col'], decision_col):
        print(c_sheet.cell(row=active_row,
                           column=active_column))
        pie_specs = fetch_expected_specs_in_(c_sheet, active_row)
        # generate a unique reference
        puid = str(uuid.uuid4())
        pie_data.update({
            puid: pie_specs})

# a debug flag for inspecting pie_data contents
pass
